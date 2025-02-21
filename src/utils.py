import inspect
# import torch
import random
# import numpy as np
import re
import openpyxl
import regex
import json
import time
import spacy
import sys

from types import MethodType, FunctionType
from pathlib import Path
from src.rfc import RFC
from src.configs.common_configs import PathConfig
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity as cos_sim
from src.extraction import run
from src.model import ModelFactory
from tqdm import tqdm


def enable_grad_for_hf_llm(func: MethodType | FunctionType) -> MethodType | FunctionType:
    return func.__closure__[1].cell_contents

def get_script_name() -> str:
    caller_frame_record = inspect.stack()[1]
    module_path = caller_frame_record.filename
    return Path(module_path).stem

def set_seed(seed: int) -> None:
    random.seed(seed)
    # np.random.seed(seed)
    # torch.manual_seed(seed)
    # if torch.cuda.is_available():
    #     torch.cuda.manual_seed(seed)
    #     torch.cuda.manual_seed_all(seed)
    #     torch.backends.cudnn.deterministic = True
    #     torch.backends.cudnn.benchmark = False

def split_document_by_sections(rfc : str) -> dict:
    """将RFC文档按章节划分。
    """
    rfc_4271_filter = ['10.  Each timer has a "timer" and a "time" (the initial value).']
    
    rfc_2328_filter = ["6.  Note that Router RT10 has a virtual link configured to Router",
                       "1.  Into the backbone, Router RT4 originates separate",
                       "8.  Two of the summary-LSAs originated by Router RT4",
                       "16.  There are three OSPF routers (RTA, RTB and RTC)",
                       "12.2.  The lookup process may return an LSA whose LS age is equal to",
                       "224.0.0.5.  All routers running OSPF should be prepared to",
                       "224.0.0.6.  Both the Designated Router and Backup Designated",
                       "10.9.  The reception of Link State Request packets is documented in",
                       "2178.  All differences are backward-compatible. Implementations of",
                       "16.2.  After completion of the calculation in Section 16.3, any"]
    
    if rfc == "4271":
        rfc_path = RFC().file_path(rfc)
        sections = {}
        current_section = None
        with open(rfc_path, 'r') as file:
            for line in file.readlines():
                line = line.strip()
                # 使用正则表达式匹配章节标题格式
                if re.match(r'(\d+\.)+\s{2}', line) and not re.match(r'.*\.{2,}.*', line) and line not in rfc_4271_filter:
                    current_section = line
                    sections[current_section] = ""
                elif current_section is not None:
                    sections[current_section] += line
        return sections
    elif rfc == "2328":
        rfc_path = RFC().file_path(rfc)
        sections = {}
        current_section = None
        with open(rfc_path, 'r') as file:
            for line in file.readlines():
                line = line.strip()
                # 使用正则表达式匹配章节标题格式
                if re.match(r'(\d+\.)+\s{2}', line) and not re.match(r'.*\.{2,}.*', line) and line not in rfc_2328_filter:
                    current_section = line
                    sections[current_section] = ""
                elif current_section is not None:
                    sections[current_section] += line
        return sections
    else:
        rfc_path = RFC().file_path(rfc)
        sections = {}
        current_section = None
        with open(rfc_path, 'r') as file:
            for line in file.readlines():
                line = line.strip()
                # 使用正则表达式匹配章节标题格式
                if re.match(r'(\d+\.)+\s{2}', line) and not re.match(r'.*\.{2,}.*', line):
                    current_section = line
                    sections[current_section] = ""
                elif current_section is not None:
                    sections[current_section] += line
        return sections

def extract_formatted_rules(log_path) -> list:
    """从日志文件中提取格式化的规则。
    """
    # 读取日志文件
    with open(log_path, 'r') as file:
        log = file.read()
        file.close()
    # 使用正则表达式匹配规则
    pattern = r'<RULE>(.*?)<\/RULE>'
    
    rules = list(set(re.findall(pattern, log)))
    return rules

def insert2excel(rfc, log_name, log_path):
    """将提取的规则插入到Excel文件中。
    Args:
        rfc (str): RFC文档的编号
        log_name (str): 日志文件名，保存的是从RFC文档中提取的规则
        log_path (str): 日志文件所在的路径
    """
    
    is_new = False
    
    # 提取格式化的规则
    rules = extract_formatted_rules(log_path / log_name)
    
    # excel保存在日志同级目录下
    excel_name = "extracted_rules.xlsx"
    excel_path = log_path / excel_name
    
    try:
        # 尝试打开已有的Excel文件
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        # 如果文件不存在，则创建新的工作簿，并删除默认的'Sheet'工作表
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
    
    table_name = f"{rfc}_PKF"
    
    try:
        worksheet = workbook[table_name]
    except KeyError:
        is_new = True
        worksheet = workbook.create_sheet(table_name)
    
    column_count = worksheet.max_column
    
    if is_new:
        current_column = 1
    else:
        current_column = column_count + 1
    
    for index, value in enumerate(rules):
        if index == 0:
            worksheet.cell(row = index + 1, column = current_column).value = log_name
            bold_font = Font(bold=True)
            worksheet.cell(row = index + 1, column = current_column).font = bold_font
        else:
            worksheet.cell(row = index + 1, column = current_column).value = value
    
    workbook.save(excel_path)

def cosine_similarity(str1, str2):
    """计算两个字符串的余弦相似度。
    """
    vectorizer = CountVectorizer().fit_transform([str1, str2])
    vectors = vectorizer.toarray()
    return cos_sim([vectors[0]], [vectors[1]])[0][0]
        
def get_class_by_name(
    class_name: str, 
    module_name: str = None
) -> type:
    """根据类名从指定模块（默认为调用者模块）中获取类对象
    """
    # 确定要搜索的模块
    if module_name is None:
        # 获取调用者模块（如 model.py）
        caller_frame = sys._getframe(1)
        caller_module = caller_frame.f_globals['__name__']
        module = sys.modules[caller_module]
    else:
        module = sys.modules[module_name]

    # 在目标模块中查找类
    target_name = class_name.lower()
    for obj in vars(module).values():
        if isinstance(obj, type) and obj.__name__.lower() == target_name:
            return obj
    raise ValueError(f"Class '{class_name}' not found in module {module.__name__}")

def find_key_in_json(
    data: dict | list,
    target_key: str,
    path: list[str | int] | None = None
) -> tuple[bool, list[str | int]]:
    """在嵌套的 JSON 数据（字典或列表）中递归查找指定键，并返回其路径。

    Args:
        data (dict | list): 要搜索的 JSON 数据。
        target_key (str): 需要查找的目标键名。
        path (list, optional): 递归使用的路径记录，调用时无需手动传递。默认为 None。

    Returns:
        tuple[bool, list]: 
            - 布尔值表示是否找到目标键。
            - 列表包含键的完整路径（由字典键和列表索引组成）。

    示例:
        >>> data = {"a": [{"b": {"target_key": 1}}]}
        >>> found, path = find_key_in_json(data, "target_key")
        >>> print(found, path)
        True ['a', 0, 'b', 'target_key']
    """
    if path is None:
        path = []
    if isinstance(data, dict):
        for key, value in data.items():
            if key == target_key:
                return True, path + [key]
            found, sub_path = find_key_in_json(value, target_key, path + [key])
            if found:
                return found, sub_path
    elif isinstance(data, list):
        for index, item in enumerate(data):
            found, sub_path = find_key_in_json(item, target_key, path + [index])
            if found:
                return found, sub_path
    return False, []

def find_value_in_json(
    data: dict | list,
    target_value: str,
    path: list[str | int] | None = None
) -> tuple[bool, list[str | int]]:
    """在嵌套的 JSON 数据（字典或列表）中递归查找指定字符串值，并返回其路径。

    函数会深度优先搜索嵌套结构，返回第一个匹配值的路径。
    注意：仅匹配字符串类型的值，其他类型（如数字、布尔值）会被忽略。

    Args:
        data (dict | list): 
            要搜索的 JSON 数据（支持嵌套的字典或列表）。
        target_value (str): 
            需要查找的目标字符串值。
        path (list[str | int], optional): 
            递归使用的路径记录，调用时无需手动传递。默认为 None。

    Returns:
        tuple[bool, list[str | int]]: 
            - 布尔值表示是否找到目标值。
            - 列表包含值的完整路径（由字典键和列表索引组成）。

    Raises:
        TypeError: 如果输入数据不是字典或列表。

    示例:
        >>> example_data = {
        ...     "a": {
        ...         "b": [1, {"c": "target"}],
        ...         "d": "not_a_target"
        ...     }
        ... }
        >>> found, path = find_value_in_json(example_data, "target")
        >>> print(found, path)
        True ['a', 'b', 1, 'c']
    """
    if path is None:
        path = []
    if isinstance(data, dict):
        for key, value in data.items():
            # 检查当前值是否是目标字符串
            if isinstance(value, str) and value == target_value:
                return True, path + [key]
            # 递归搜索嵌套结构
            found, sub_path = find_value_in_json(value, target_value, path + [key])
            if found:
                return found, sub_path
    elif isinstance(data, list):
        for index, item in enumerate(data):
            # 检查当前元素是否是目标字符串
            if isinstance(item, str) and item == target_value:
                return True, path + [index]
            # 递归搜索嵌套结构
            found, sub_path = find_value_in_json(item, target_value, path + [index])
            if found:
                return found, sub_path
    return False, []

def insert_into_json(
    data: dict | list,
    path: list[str | int],
    new_data: dict
) -> None:
    """根据指定路径在嵌套的 JSON 数据（字典或列表）中插入或更新数据。

    函数会直接修改原始数据（原地操作），不返回新对象。

    Args:
        data (dict | list): 要修改的 JSON 数据（支持嵌套的字典或列表）。
        path (list[str | int]): 目标路径，由字典键（字符串）和列表索引（整数）组成。
            示例：`["a", 0, "b"]` 表示 `data["a"][0]["b"]`。
        new_data (dict): 要插入或更新的字典数据。

    Raises:
        TypeError: 如果路径中的索引类型与数据结构不匹配（例如对字典使用整数索引）。
        IndexError: 如果路径中的列表索引超出范围。

    示例:
        >>> data = {"a": [{"b": 1}]}
        >>> insert_into_json(data, ["a", 0], {"c": 2})
        >>> print(data)
        {'a': [{'b': 1, 'c': 2}]}
    """
    def helper(sub_data: dict | list, sub_path: list[str | int]) -> None:
        # 处理路径为空的情况（直接合并到当前层级）
        if not sub_path:
            if isinstance(sub_data, dict):
                sub_data.update(new_data)
            else:
                raise TypeError("路径为空时，顶层数据结构必须是字典")
            return
        
        if len(sub_path) == 1:
            current_key = sub_path[0]
            if isinstance(sub_data, dict):
                if current_key in sub_data:
                    # 合并或覆盖字典
                    if isinstance(sub_data[current_key], dict):
                        sub_data[current_key].update(new_data)
                    else:
                        sub_data[current_key] = {**new_data}
                else:
                    # 创建新键并插入数据
                    sub_data[current_key] = new_data
            elif isinstance(sub_data, list) and isinstance(current_key, int):
                if current_key < len(sub_data):
                    # 合并或覆盖列表元素
                    if isinstance(sub_data[current_key], dict):
                        sub_data[current_key].update(new_data)
                    else:
                        sub_data[current_key] = {**new_data}
                else:
                    raise IndexError(f"列表索引 {current_key} 超出范围")
        else:
            current_key = sub_path[0]
            if isinstance(sub_data, dict):
                if current_key in sub_data:
                    helper(sub_data[current_key], sub_path[1:])
                else:
                    raise KeyError(f"键 '{current_key}' 不存在于字典中")
            elif isinstance(sub_data, list) and isinstance(current_key, int):
                if current_key < len(sub_data):
                    helper(sub_data[current_key], sub_path[1:])
                else:
                    raise IndexError(f"列表索引 {current_key} 超出范围")
            else:
                raise TypeError("路径类型与数据结构不匹配")

    helper(data, path)

def delete_key_in_json(data: dict | list, key: str) -> None:
    """在嵌套的 JSON 数据（字典或列表）中递归删除所有指定的键及其对应的值。

    函数会直接修改原始数据（原地操作），不返回任何值。
    注意：若目标键在嵌套结构中出现多次，所有实例均会被删除。

    Args:
        data (dict | list): 要修改的 JSON 数据（支持嵌套的字典或列表）。
        key (str): 需要删除的目标键名。

    Raises:
        TypeError: 如果输入数据不是字典或列表。

    示例:
        >>> data = {"a": {"key": "value", "b": [{"key": 1}, 2]}}
        >>> delete_key_in_json(data, "key")
        >>> print(data)
        {'a': {'b': [{}, 2]}}
    """
    if isinstance(data, dict):
        # 遍历字典的键（转换为列表避免修改冲突）
        for k in list(data.keys()):
            if k == key:
                del data[k]
            else:
                # 仅当值是字典或列表时递归
                value = data[k]
                if isinstance(value, (dict, list)):
                    delete_key_in_json(value, key)
    elif isinstance(data, list):
        # 遍历列表元素，仅当元素是字典或列表时递归
        for item in data:
            if isinstance(item, (dict, list)):
                delete_key_in_json(item, key)
    else:
        raise TypeError("输入数据必须是字典或列表")




